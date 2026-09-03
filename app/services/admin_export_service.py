"""管理后台取证导出服务（v10 变更 D，R2 §5.1/§5.4）。

把原先散在 `app/routers/admin.py` 的 `_build_conversation` / `_build_forensic_row`
下沉到服务层，并扩展出：

- 结构化对话 `build_conversation_rows`（详情接口用）；
- `scope=profile` 的精简取证行 `build_profile_row`；
- **统一长表** `build_long_rows`：每条记录的每个字段展开成一行
  `[记录ID, 字段, 值, 说明]`，**整张文件就是一张表**，不再有"单独一块字段说明"；
- **对话流水** `build_transcripts`：匹配记录下的双方对话，抽离出长表之外，
  以「时间 角色：内容」的可读形式单独呈现（满足"完整说话流程 + 时间戳"诉求）；
- 三种格式渲染器 `render_csv` / `render_xlsx` / `render_md`：
  - 长表一张（记录ID|字段|值|说明）；
  - 对话流水一块（独立于长表，可读时间线）；
  - 「说明」列对编码字段（如匹配状态 2、操作类型 handover_complete）**直接解出
    该值的具体含义**，不再只写字段名。

设计约束：

1. **openpyxl 惰性导入**：模块导入期绝不 `import openpyxl`。未安装时
   `render_xlsx` 抛 `ExportDependencyError`，路由层转 400 + `code 9001`，
   **禁止 500 堆栈**（AC-D12）。
2. **md 零依赖**：纯 f-string 拼装，内容需转义 `|` 与换行，否则表格被撑破。
3. `build_conversation` 的输出格式与 v7 **逐字节一致**（`[iso] 角色: 内容`，
   以 ` ⏎ ` 连接），`test_v7_admin_export.py` 的对话文本断言依赖该格式。
4. 长表行一律 `[记录ID, 字段, 值, 说明]` 四列：匹配记录行用 `match.id` 作记录 ID。
5. 对话流水**不在长表内**，分 `scope` 决定：
   - `profile`：仅长表；
   - `conversation`：仅对话流水；
   - `all`：长表 + 对话流水。
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Iterable

from sqlalchemy.orm import Session

from app.models.im import IMMessage, IMSession
from app.models.item import FoundItem, LostItem
from app.models.match import MatchRecord
from app.models.user import User

# ---------------------------------------------------------------------------
# 字段定义（scope → 列顺序）。仍保留宽表常量供 build_profile_row /
# build_conversation_export_rows 复用；新格式的列定义见 LONG_FORMAT_COLUMNS。
# ---------------------------------------------------------------------------
FORENSIC_FIELDS: list[str] = [
    "match_id",
    "lost_item_id", "lost_category", "lost_title", "lost_description",
    "lost_images", "lost_student_no", "lost_phone",
    "found_item_id", "found_category", "found_description",
    "found_images", "found_student_no", "found_phone",
    "completed_at",
    "conversation",
]

PROFILE_FIELDS: list[str] = [
    "match_id",
    "lost_item_id", "lost_category", "lost_title",
    "lost_student_no", "lost_phone", "lost_real_name",
    "found_item_id", "found_category",
    "found_student_no", "found_phone", "found_real_name",
    "match_score", "status", "completed_at",
]

CONVERSATION_FIELDS: list[str] = ["match_id", "message_id", "sent_at", "role_label", "content"]

# scope → (列定义, Sheet 名)；保留以便旧调用方（如 build_profile_row）复用。
SCOPE_FIELDS: dict[str, list[str]] = {
    "profile": PROFILE_FIELDS,
    "conversation": CONVERSATION_FIELDS,
    "all": FORENSIC_FIELDS,
}

# ---------------------------------------------------------------------------
# 长表列定义（一张表四个列）。所有导出格式都按这四列输出，没有额外的字段说明块。
# ---------------------------------------------------------------------------
LONG_FORMAT_COLUMNS: list[str] = ["记录ID", "字段", "值", "说明"]

# ---------------------------------------------------------------------------
# 字段含义表（字段 → 基础含义）。导出时「说明」列 = 基础含义 +（编码字段）该值的具体含义。
# ---------------------------------------------------------------------------
MATCH_FIELD_MEANING: dict[str, str] = {
    "match_id": "匹配记录ID（系统内部编号，关联失物与拾物）",
    "lost_item_id": "失物条目ID",
    "lost_category": "失物类别（如 手机/钱包/钥匙/书包）",
    "lost_title": "失物主标题",
    "lost_description": "失物文字描述",
    "lost_images": "失物图片URL，多张以 | 分隔",
    "lost_student_no": "失主学号（明文，仅管理员可见）",
    "lost_phone": "失主手机号（明文，仅管理员可见）",
    "lost_real_name": "失主真实姓名（明文）",
    "found_item_id": "拾物条目ID",
    "found_category": "拾物类别（中文）",
    "found_description": "拾物文字描述",
    "found_images": "拾物图片URL，多张以 | 分隔",
    "found_student_no": "拾主学号（明文，仅管理员可见）",
    "found_phone": "拾主手机号（明文，仅管理员可见）",
    "found_real_name": "拾主真实姓名（明文）",
    "match_score": "匹配综合得分（0-100，≥80 为疑似匹配阈值）",
    "status": "匹配状态",
    "completed_at": "交接完成时间（空=尚未完成交接）",
    "conversation": "该匹配下全部对话（时间 角色：内容，按时间排序）",
    "message_id": "IM 消息记录ID",
    "sent_at": "消息发送时间",
    "role_label": "发送者角色",
    "content": "消息正文",
}

# ---------------------------------------------------------------------------
# 编码字段取值字典：让「说明」列直接解出具体含义，而不是只写字段名。
# ---------------------------------------------------------------------------
# 匹配状态（来源：app/models/match.py MatchRecord.status 注释）
MATCH_STATUS: dict[int, str] = {
    0: "待认领（失主尚未确认认领）",
    1: "认领中（失主已申请认领，等待交接）",
    2: "已完成（双方完成线下交接，匹配闭环）",
    3: "已拒绝（失主拒绝该认领）",
    4: "待自取（物品待失主自行取走）",
    5: "已放弃（认领方放弃）",
    6: "已撤回（认领被撤回）",
}
_MATCH_STATUS_SCALE = "0待认领/1认领中/2已完成/3已拒绝/4待自取/5已放弃/6已撤回"

_ROLE_LABELS: dict[int, str] = {0: "失主", 1: "拾得者"}

_MEDIA_TYPES: dict[str, str] = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "md": "text/markdown; charset=utf-8",
}


class ExportDependencyError(RuntimeError):
    """导出所需的可选依赖缺失（当前仅 xlsx 的 openpyxl）。

    由路由层捕获并转成 `400 + {"code": 9001, "message": ...}`，绝不冒泡成 500。
    """


def role_label(sender_role) -> str:
    """把 `IMMessage.sender_role` 映射为中文角色名（未知值按「拾得者」兜底）。"""
    try:
        return _ROLE_LABELS.get(int(sender_role), "拾得者")
    except (TypeError, ValueError):
        return "拾得者"


def _explain_match_field(field: str, value) -> str:
    """字段「说明」：基础含义 +（编码字段）该值的具体含义。

    - `status=2` → "匹配状态：2=已完成（双方完成线下交接，匹配闭环）（取值：0待认领/.../6已撤回）"
    - `match_score=71.0` → "匹配综合得分（0-100，≥80 为疑似匹配阈值）：71.0"
    - 普通字段（如 lost_title） → 仅基础含义
    """
    base = MATCH_FIELD_MEANING.get(field, "")
    if field == "status":
        try:
            code = int(value)
        except (TypeError, ValueError):
            return base
        meaning = MATCH_STATUS.get(code, "未知状态")
        return f"匹配状态：{code}={meaning}（全部取值：{_MATCH_STATUS_SCALE}）"
    if field == "match_score":
        return f"匹配综合得分（0-100，≥80 为疑似匹配阈值）：{value}"
    return base


# ---------------------------------------------------------------------------
# 对话
# ---------------------------------------------------------------------------
def _load_messages(db: Session, match_id: int) -> list[IMMessage]:
    """取该匹配下全部 IM 消息（按 sent_at 升序）；无会话返回空列表。"""
    session_ids = [
        sid for (sid,) in db.query(IMSession.id)
        .filter(IMSession.match_id == match_id).all()
    ]
    if not session_ids:
        return []
    return (
        db.query(IMMessage)
        .filter(IMMessage.session_id.in_(session_ids))
        .order_by(IMMessage.sent_at.asc())
        .all()
    )


def build_conversation(db: Session, match_id: int) -> str:
    """拼接该匹配下全部 IM 对话文本（按 sent_at 升序），单行以 ⏎ 分隔。

    ⚠️ 输出格式与 v7 完全一致，存量测试依赖之，不得修改。

    Args:
        db: 数据库会话。
        match_id: 匹配 id。

    Returns:
        `"[iso] 失主: 内容 ⏎ [iso] 拾得者: 内容"`；无会话/无消息返回空串。
    """
    msgs = _load_messages(db, match_id)
    if not msgs:
        return ""
    parts = []
    for m in msgs:
        ts = m.sent_at.isoformat() if m.sent_at else ""
        parts.append(f"[{ts}] {role_label(m.sender_role)}: {m.content}")
    return " ⏎ ".join(parts)


def build_conversation_rows(db: Session, match_id: int) -> list[dict]:
    """结构化对话（详情接口用）。

    Args:
        db: 数据库会话。
        match_id: 匹配 id。

    Returns:
        `[{"id": int, "sent_at": datetime|None, "sender_role": int, "role_label": str,
        "content": str}, ...]`，按 sent_at 升序；无会话返回 `[]`。
    """
    rows: list[dict] = []
    for m in _load_messages(db, match_id):
        try:
            sender_role = int(m.sender_role)
        except (TypeError, ValueError):
            sender_role = 1
        rows.append({
            "id": m.id,
            "sent_at": m.sent_at,
            "sender_role": sender_role,
            "role_label": role_label(m.sender_role),
            "content": str(m.content or ""),
        })
    return rows


def _format_message_time(sent_at) -> str:
    """消息时间 → `YYYY-MM-DD HH:MM:SS`（无时区标记，与存储值一致）。"""
    if not sent_at:
        return "（时间未知）"
    if isinstance(sent_at, str):
        return sent_at
    return sent_at.strftime("%Y-%m-%d %H:%M:%S")


def format_conversation_transcript(db: Session, match_id: int) -> str:
    """对话可读流水：`时间  角色：内容` 逐行，按时间升序。

    用于导出文件里「独立于长表」的对话区块，呈现双方完整说话流程 + 时间戳。
    """
    msgs = _load_messages(db, match_id)
    if not msgs:
        return "（该匹配无对话记录）"
    lines = []
    for m in msgs:
        ts = _format_message_time(m.sent_at)
        lines.append(f"{ts}  {role_label(m.sender_role)}：{m.content}")
    return "\n".join(lines)


def build_transcripts(db: Session, matches: Iterable[MatchRecord], scope: str) -> list[dict]:
    """构建对话流水（独立于长表）。

    Args:
        db: 数据库会话。
        matches: 匹配记录序列。
        scope: `profile`（不产出对话）/ `conversation` / `all`（产出对话）。

    Returns:
        `[{"match_id": int, "messages": [{"time": str, "role": str, "content": str}, ...]}, ...]`
        `profile` scope 返回 `[]`。
    """
    if scope == "profile":
        return []
    out: list[dict] = []
    for match in matches:
        msgs = _load_messages(db, match.id)
        messages = [
            {
                "time": _format_message_time(m.sent_at),
                "role": role_label(m.sender_role),
                "content": str(m.content or ""),
            }
            for m in msgs
        ]
        out.append({"match_id": match.id, "messages": messages})
    return out


# ---------------------------------------------------------------------------
# 取证行（宽表，保留兼容）
# ---------------------------------------------------------------------------
def _sides(db: Session, match: MatchRecord):
    """取匹配双方的物品与用户对象（任一缺失以 None 返回，不抛异常）。"""
    lost = db.get(LostItem, match.lost_id)
    found = db.get(FoundItem, match.found_id)
    lost_user = db.get(User, lost.publisher_id) if lost else None
    found_user = db.get(User, found.finder_id) if found else None
    return lost, found, lost_user, found_user


def build_forensic_row(db: Session, match: MatchRecord) -> dict:
    """构建单条全量取证行（含双方明文 student_no/phone、扁平对话文本）。

    列集合 = `FORENSIC_FIELDS`，与 v7 完全一致。保留以兼容 `_build_forensic_row` 别名。
    """
    lost, found, lost_user, found_user = _sides(db, match)
    return {
        "match_id": match.id,
        "lost_item_id": lost.id if lost else "",
        "lost_category": lost.category_name if lost else "",
        "lost_title": lost.title if lost else "",
        "lost_description": lost.description if lost else "",
        "lost_images": "|".join(lost.images or []) if lost and lost.images else "",
        "lost_student_no": lost_user.student_no if lost_user else "",
        "lost_phone": lost_user.phone if lost_user else "",
        "found_item_id": found.id if found else "",
        "found_category": found.category_name if found else "",
        "found_description": found.description if found else "",
        "found_images": "|".join(found.images or []) if found and found.images else "",
        "found_student_no": found_user.student_no if found_user else "",
        "found_phone": found_user.phone if found_user else "",
        "completed_at": match.completed_at.isoformat() if match.completed_at else "",
        "conversation": build_conversation(db, match.id),
    }


def build_profile_row(db: Session, match: MatchRecord) -> dict:
    """构建单条「个人信息」行（`scope=profile`）：双方身份 + 物品摘要，**不含对话**。

    列集合 = `PROFILE_FIELDS`。
    """
    lost, found, lost_user, found_user = _sides(db, match)
    return {
        "match_id": match.id,
        "lost_item_id": lost.id if lost else "",
        "lost_category": lost.category_name if lost else "",
        "lost_title": lost.title if lost else "",
        "lost_student_no": lost_user.student_no if lost_user else "",
        "lost_phone": lost_user.phone if lost_user else "",
        "lost_real_name": (lost_user.real_name or "") if lost_user else "",
        "found_item_id": found.id if found else "",
        "found_category": found.category_name if found else "",
        "found_student_no": found_user.student_no if found_user else "",
        "found_phone": found_user.phone if found_user else "",
        "found_real_name": (found_user.real_name or "") if found_user else "",
        "match_score": float(match.match_score) if match.match_score is not None else "",
        "status": int(match.status) if match.status is not None else "",
        "completed_at": match.completed_at.isoformat() if match.completed_at else "",
    }


# ---------------------------------------------------------------------------
# 长表：匹配记录字段（不含对话），所有导出格式的唯一表数据源
# ---------------------------------------------------------------------------
def build_long_rows(db: Session, matches: Iterable[MatchRecord], scope: str) -> list[dict]:
    """构建长表行：每条匹配记录的每个字段一行，列 = `[记录ID, 字段, 值, 说明]`。

    - `profile` / `all`：输出匹配记录字段（PROFILE_FIELDS）。
    - `conversation`：**不输出长表行**（对话由 `build_transcripts` 单独呈现）。

    「说明」列对编码字段直接解出该值的具体含义（见 `_explain_match_field`）。

    Args:
        db: 数据库会话。
        matches: 匹配记录序列（调用方已过滤掉不存在的 id）。
        scope: `profile` / `conversation` / `all`。

    Returns:
        行 dict 列表，键集合与 `LONG_FORMAT_COLUMNS` 一致。
    """
    rows: list[dict] = []
    if scope in ("profile", "all"):
        for match in matches:
            profile = build_profile_row(db, match)
            for field in PROFILE_FIELDS:
                value = profile.get(field, "")
                rows.append({
                    "记录ID": match.id,
                    "字段": field,
                    "值": value,
                    "说明": _explain_match_field(field, value),
                })
    return rows


# ---------------------------------------------------------------------------
# 渲染器：长表一张 + 对话流水一块
# ---------------------------------------------------------------------------
def export_filename(scope: str, ext: str) -> str:
    """导出文件名 `forensic_matches_{scope}_{YYYYMMDD}.{ext}`。"""
    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    return f"forensic_matches_{scope}_{date_str}.{ext}"


_LEGEND_HEADER = (
    "本文件由失物招领系统自动生成，记录与审计日志一致，"
    "可作为溯源与责任认定（追责）依据。时间均为 UTC（ISO8601）。"
)


def _transcript_block(transcripts: list[dict]) -> list[str]:
    """对话流水 → 注释行列表（每条消息一行 `时间  角色：内容`）。

    用于 CSV / MD 里独立于长表的对话区块。
    """
    lines: list[str] = []
    for t in transcripts:
        msgs = t.get("messages", [])
        lines.append(f"# ===== 对话记录（匹配 #{t['match_id']}，共 {len(msgs)} 条）=====")
        if not msgs:
            lines.append("# （该匹配无对话记录）")
            continue
        for m in msgs:
            lines.append(f"# {m['time']}  {m['role']}：{m['content']}")
    return lines


def render_csv(rows: list[dict], transcripts: list[dict], scope: str) -> tuple[bytes, str, str]:
    """渲染 CSV（UTF-8）：**长表一张** `[记录ID, 字段, 值, 说明]` + **对话流水一块**。

    - 顶部第一行 `#` 注释写明取证声明；
    - 接着是长表（scope=conversation 时无长表，直接进对话区块）；
    - 长表之后是 `#` 注释形式的对话流水，独立于长表，呈现双方完整说话流程 + 时间戳。

    Returns:
        `(内容字节, media_type, 文件名)`。`filename` 由调用方 `render()` 按 scope 决定。
    """
    buf = io.StringIO()
    buf.write(f"# {_LEGEND_HEADER}\n")
    if rows:
        writer = csv.DictWriter(buf, fieldnames=LONG_FORMAT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in LONG_FORMAT_COLUMNS})
        buf.write("\n")
    if transcripts:
        for line in _transcript_block(transcripts):
            buf.write(line + "\n")
        buf.write("\n")
    return buf.getvalue().encode("utf-8"), _MEDIA_TYPES["csv"]


def _autofit_long(ws, rows: list[dict]) -> None:
    """按该列最长值近似自适应列宽（上限 60）。"""
    from openpyxl.utils import get_column_letter

    for idx, col in enumerate(LONG_FORMAT_COLUMNS, start=1):
        longest = len(str(col))
        for row in rows:
            longest = max(longest, len(str(row.get(col, ""))))
        ws.column_dimensions[get_column_letter(idx)].width = min(longest + 2, 60)


def render_xlsx(rows: list[dict], transcripts: list[dict], scope: str) -> tuple[bytes, str, str]:
    """渲染 xlsx（openpyxl **惰性导入**）：

    - Sheet「取证记录」：整张长表 `[记录ID, 字段, 值, 说明]`（scope=conversation 时为空表，仅表头）；
    - Sheet「对话记录」：独立于长表，列 = `[匹配ID, 时间, 发送方, 内容]`，逐条呈现双方对话。

    Raises:
        ExportDependencyError: 未安装 openpyxl（路由层转 400，禁止 500）。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:      # pragma: no cover - 依赖可用性由部署环境决定
        raise ExportDependencyError("服务器未安装 openpyxl，无法导出 xlsx") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "取证记录"
    ws.append(LONG_FORMAT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append([str(row.get(k, "")) for k in LONG_FORMAT_COLUMNS])
    _autofit_long(ws, rows)

    # 对话记录（独立 Sheet）
    ws2 = wb.create_sheet(title="对话记录")
    conv_header = ["匹配ID", "时间", "发送方", "内容"]
    ws2.append(conv_header)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    for t in transcripts:
        for m in t.get("messages", []):
            ws2.append([t["match_id"], m["time"], m["role"], m["content"]])
    for idx in range(1, 5):
        ws2.column_dimensions[get_column_letter(idx)].width = 24 if idx != 4 else 50

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), _MEDIA_TYPES["xlsx"]


def _md_escape(value) -> str:
    """转义 Markdown 表格单元格内容（`|` 与换行会撑破表格）。"""
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def render_md(
    rows: list[dict], transcripts: list[dict], scope: str, match_count: int, exported_by: int | None = None
) -> tuple[bytes, str, str]:
    """渲染 Markdown：**长表一张** + **对话流水一块**（独立 `## 对话记录` 段落）。

    Returns:
        `(内容字节, media_type, 文件名)`。`filename` 由调用方 `render()` 按 scope 决定。
    """
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    lines: list[str] = [
        "# 失物招领取证导出",
        "",
        f"- 导出时间：{date_str}",
        f"- 导出范围：{scope}",
        f"- 匹配条数：{match_count}",
        f"- 长表行数：{len(rows)}",
    ]
    if exported_by is not None:
        lines.append(f"- 导出人（管理员ID）：{exported_by}")
    lines += [
        "",
        f"> {_LEGEND_HEADER}",
        "",
    ]
    if rows:
        lines += [
            "| 记录ID | 字段 | 值 | 说明 |",
            "| --- | --- | --- | --- |",
        ]
        for row in rows:
            lines.append(
                f"| {row.get('记录ID', '')} | "
                f"{_md_escape(row.get('字段', ''))} | "
                f"{_md_escape(row.get('值', ''))} | "
                f"{_md_escape(row.get('说明', ''))} |"
            )
        lines.append("")

    if transcripts:
        lines.append("## 对话记录")
        lines.append("")
        for t in transcripts:
            msgs = t.get("messages", [])
            lines.append(f"### 匹配 #{t['match_id']}（共 {len(msgs)} 条）")
            lines.append("")
            if not msgs:
                lines.append("（该匹配无对话记录）")
                lines.append("")
                continue
            for m in msgs:
                lines.append(f"{m['time']}  {m['role']}：{m['content']}")
            lines.append("")

    content = "\n".join(lines) + "\n"
    return content.encode("utf-8"), _MEDIA_TYPES["md"]


def render(
    db: Session, matches: list[MatchRecord], scope: str, fmt: str, exported_by: int | None = None
) -> tuple[bytes, str, str]:
    """按 `(scope, format)` 分派到对应渲染器。

    所有渲染器现在只输出**一张长表** `[记录ID, 字段, 值, 说明]` + **一块对话流水**
    （对话独立于长表，呈现双方完整说话流程 + 时间戳）。「说明」列对编码字段直接解出
    该值的具体含义。

    Args:
        db: 数据库会话。
        matches: 匹配记录列表。
        scope: `profile` / `conversation` / `all`（调用方须已校验）。
        fmt: `csv` / `xlsx` / `md`（调用方须已校验）。
        exported_by: 触发导出的管理员用户 ID（accountability 留痕，可空）。

    Returns:
        `(内容字节, media_type, 文件名)`。

    Raises:
        ExportDependencyError: xlsx 且 openpyxl 缺失。
    """
    rows = build_long_rows(db, matches, scope)
    transcripts = build_transcripts(db, matches, scope)
    filename = export_filename(scope, fmt)
    if fmt == "xlsx":
        content, media_type = render_xlsx(rows, transcripts, scope)
    elif fmt == "md":
        content, media_type = render_md(rows, transcripts, scope, len(matches), exported_by)
    else:
        content, media_type = render_csv(rows, transcripts, scope)
    return content, media_type, filename
